from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook, Workbook
from io import BytesIO
from datetime import datetime
import logging

from services.database import get_db
from models.catalog_items import CatalogItem
from models.user_generations import UserGeneration
from models.log import Log
from models.users import User
from services.auth import get_current_admin_user, get_current_active_user

router = APIRouter(prefix="/excel", tags=["Excel"])

# Модульный логгер
logger = logging.getLogger(__name__)

@router.post("/upload-items")
async def upload_items_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Загрузка товаров из Excel в ОБЩИЙ КАТАЛОГ (только для админов)
    
    Обязательные колонки в Excel:
    - id_item (ID товара)
    - name (название)
    - price (цена)
    - photoUrl (URL фото)
    - slug (SEO-slug)
    
    Опциональные колонки:
    - stuff (материал)
    - category_id
    - raw_description (описание)
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате Excel (.xlsx или .xls)")
    
    try:
        logger.info("[UPLOAD] Начало загрузки файла: %s", file.filename)
        contents = await file.read()
        logger.debug("[UPLOAD] Файл прочитан, размер: %d байт", len(contents))
        
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
        logger.info("[UPLOAD] Excel файл открыт, лист: %s", ws.title)
        
        # Читаем заголовки (первая строка)
        headers = [cell.value for cell in ws[1]]
        logger.debug("[UPLOAD] Заголовки найдены: %s", headers)
        
        # Проверяем обязательные поля
        required = ['id_item', 'name', 'price', 'photoUrl', 'slug']
        missing = [field for field in required if field not in headers]
        if missing:
            logger.error("[UPLOAD] Ошибка: отсутствуют колонки: %s", missing)
            raise HTTPException(
                status_code=400,
                detail=f"Отсутствуют обязательные колонки: {', '.join(missing)}"
            )
        
        # Создаём мапинг индексов колонок
        col_map = {header: idx for idx, header in enumerate(headers)}
        
        added_count = 0
        skipped_count = 0
        errors = []
        
        # Обрабатываем строки (пропускаем заголовок)
        logger.info("[UPLOAD] Начинаем обработку строк...")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Извлекаем данные
                id_item = row[col_map['id_item']]
                name = row[col_map['name']]
                price = row[col_map['price']]
                photo_url = row[col_map['photoUrl']]
                slug = row[col_map['slug']]
                
                logger.debug("[UPLOAD] Строка %d: id_item=%s, name=%s, price=%s", row_idx, id_item, name, price)
                
                # Валидация
                if not id_item or not name or not price or not photo_url or not slug:
                    errors.append(f"Строка {row_idx}: пропущены обязательные поля")
                    skipped_count += 1
                    logger.warning("[UPLOAD] Строка %d: пропущена - нет обязательных полей", row_idx)
                    continue
                
                # Проверяем, нет ли уже такого товара в каталоге
                existing = await db.execute(
                    select(CatalogItem).where(CatalogItem.id_item == str(id_item))
                )
                if existing.scalar_one_or_none():
                    errors.append(f"Строка {row_idx}: товар {id_item} уже есть в каталоге")
                    skipped_count += 1
                    logger.info("[UPLOAD] Строка %d: пропущена - товар уже существует: %s", row_idx, id_item)
                    continue
                
                # Создаём товар в каталоге
                new_item = CatalogItem(
                    id_item=str(id_item),
                    name=str(name),
                    price=float(price),
                    photoUrl=str(photo_url),
                    slug=str(slug),
                    raw_description=str(row[col_map['raw_description']]) if 'raw_description' in col_map and row[col_map['raw_description']] else None,
                    stuff=str(row[col_map['stuff']]) if 'stuff' in col_map and row[col_map['stuff']] else None,
                    category_id=str(row[col_map['category_id']]) if 'category_id' in col_map and row[col_map['category_id']] else None,
                    balance=int(row[col_map['balance']]) if 'balance' in col_map and row[col_map['balance']] else 0,
                )
                
                db.add(new_item)
                added_count += 1
                logger.info("[UPLOAD] Строка %d: добавлен товар id_item=%s", row_idx, id_item)
                
            except Exception as e:
                error_msg = f"Строка {row_idx}: {str(e)}"
                errors.append(error_msg)
                skipped_count += 1
                logger.exception("[UPLOAD] Строка %d: ошибка при обработке: %s", row_idx, str(e))
        
        # Сохраняем всё в БД
        logger.info("[UPLOAD] Сохраняем в БД: %d товаров", added_count)
        await db.commit()
        logger.info("[UPLOAD] Успешно сохранено")
        
        # Логируем массовую загрузку
        log = Log(
            user_id=current_admin.id,
            action='bulk_upload_catalog',
            message=f"Загружено {added_count} товаров в каталог из Excel. Пропущено: {skipped_count}",
            status='completed' if added_count > 0 else 'error'
        )
        db.add(log)
        await db.commit()
        
        logger.info("[UPLOAD] Результат: добавлено=%d, пропущено=%d, ошибок=%d", added_count, skipped_count, len(errors))
        
        return {
            "success": True,
            "added": added_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []  # Возвращаем только первые 10 ошибок
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[UPLOAD] КРИТИЧЕСКАЯ ОШИБКА при загрузке Excel: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Ошибка обработки файла: {str(e)}")


@router.get("/backup-database")
async def backup_database_to_excel(
    db: AsyncSession = Depends(get_db),
    current_admin: User = Depends(get_current_admin_user)
):
    """
    Создание полного бэкапа базы данных в Excel (только для админов)
    
    Создаёт файл с листами:
    - CatalogItems (общий каталог)
    - UserGenerations (AI-генерации)
    - Logs (логи)
    - Users (пользователи, без паролей)
    """
    try:
        logger.info("[BACKUP] Начало создания бэкапа базы данных (admin=%s)", current_admin.email)
        wb = Workbook()
        
        # === ЛИСТ 1: Каталог товаров ===
        ws_catalog = wb.active
        ws_catalog.title = "CatalogItems"
        
        catalog_headers = [
            'id', 'id_item', 'uid', 'sid', 'name', 'slug', 'price', 'balance',
            'stuff', 'category_id', 'photoUrl', 'image_title', 'raw_description',
            'created_at', 'updated_at'
        ]
        ws_catalog.append(catalog_headers)
        
        catalog_result = await db.execute(select(CatalogItem))
        catalog_items = catalog_result.scalars().all()
        
        for item in catalog_items:
            ws_catalog.append([
                item.id, item.id_item, item.uid, item.sid, item.name, item.slug,
                item.price, item.balance, item.stuff, item.category_id,
                item.photoUrl, item.image_title, item.raw_description,
                str(item.created_at), str(item.updated_at)
            ])
        
        # === ЛИСТ 2: AI-генерации ===
        ws_generations = wb.create_sheet("UserGenerations")
        gen_headers = [
            'id', 'user_id', 'catalog_item_id', 'generation_name',
            'ai_description', 'ai_keywords', 'ai_prompt_version',
            'excel_exported', 'export_count', 'created_at', 'updated_at'
        ]
        ws_generations.append(gen_headers)
        
        gen_result = await db.execute(select(UserGeneration))
        generations = gen_result.scalars().all()
        
        for gen in generations:
            ws_generations.append([
                gen.id, gen.user_id, gen.catalog_item_id, gen.generation_name,
                gen.ai_description, gen.ai_keywords, gen.ai_prompt_version,
                gen.excel_exported, gen.export_count,
                str(gen.created_at), str(gen.updated_at)
            ])
        
        # === ЛИСТ 3: Логи ===
        ws_logs = wb.create_sheet("Logs")
        log_headers = ['id', 'user_id', 'timestamp', 'action', 'item_id', 'message', 'status']
        ws_logs.append(log_headers)
        
        logs_result = await db.execute(select(Log).order_by(Log.timestamp.desc()).limit(1000))
        logs = logs_result.scalars().all()
        
        for log in logs:
            ws_logs.append([
                log.id, log.user_id, str(log.timestamp), log.action, log.item_id, log.message, log.status
            ])
        
        # === ЛИСТ 4: Пользователи (БЕЗ паролей) ===
        ws_users = wb.create_sheet("Users")
        user_headers = ['id', 'email', 'full_name', 'role', 'is_active', 'created_at']
        ws_users.append(user_headers)
        
        users_result = await db.execute(select(User))
        users = users_result.scalars().all()
        
        for user in users:
            ws_users.append([
                user.id, user.email, user.full_name, 
                user.role.value, user.is_active, str(user.created_at)
            ])
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Логируем бэкап
        log = Log(
            user_id=current_admin.id,
            action='database_backup',
            message=f"Создан бэкап БД: {len(catalog_items)} товаров в каталоге, {len(generations)} генераций, {len(logs)} логов, {len(users)} пользователей",
            status='completed'
        )
        db.add(log)
        await db.commit()
        
        # Возвращаем файл
        filename = f"itemgate_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.exception("[BACKUP] Ошибка при создании бэкапа: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Ошибка создания бэкапа: {str(e)}")


@router.get("/export-items")
async def export_items_to_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Экспорт AI-генераций ТЕКУЩЕГО пользователя с данными о товарах в Excel
    """
    try:
        logger.info("[EXPORT] Начало экспорта генераций для user_id=%s", current_user.id)
        wb = Workbook()
        ws = wb.active
        ws.title = "My Generations"
        
        # Заголовки
        headers = [
            'ID', 'Название товара', 'Вариант генерации', 'Цена', 'URL Фото', 
            'Категория', 'Материал',
            'Описание от ИИ', 'Ключевые слова', 'Версия промпта'
        ]
        ws.append(headers)
        
        # Получаем генерации текущего пользователя с данными о товарах
        result = await db.execute(
            select(UserGeneration)
            .where(UserGeneration.user_id == current_user.id)
            .options(selectinload(UserGeneration.catalog_item))
            .order_by(UserGeneration.created_at.desc())
        )
        generations = result.scalars().all()
        
        if not generations:
            # Если нет генераций, возвращаем пустой файл
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"items_export_empty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        for gen in generations:
            item = gen.catalog_item
            
            # Безопасное извлечение данных товара
            if item:
                item_name = item.name
                item_price = item.price
                item_photo = item.photoUrl or ''
                item_category = item.category_id or '—'
                item_stuff = item.stuff or '—'
            else:
                item_name = 'ТОВАР УДАЛЁН'
                item_price = 0
                item_photo = ''
                item_category = '—'
                item_stuff = '—'
            
            ws.append([
                gen.id,
                item_name,
                gen.generation_name or 'Основной вариант',
                item_price,
                item_photo,
                item_category,
                item_stuff,
                gen.ai_description or '',
                gen.ai_keywords or '',
                gen.ai_prompt_version or ''
            ])
            
            # Обновляем счётчик экспортов
            gen.export_count += 1
            gen.excel_exported = 'exported'
        
        await db.commit()
        
        # Логируем экспорт
        log = Log(
            user_id=current_user.id,
            action='export_generations',
            message=f"Экспортировано {len(generations)} генераций в Excel",
            status='completed'
        )
        db.add(log)
        await db.commit()
        
        # Сохраняем
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"items_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.exception("[EXPORT] Ошибка при экспорте генераций: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Ошибка экспорта: {str(e)}")
